import flask
import datetime
import io
import os
import math

from openpyxl import Workbook
from openpyxl.chart import LineChart, PieChart, DoughnutChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font
from flask_socketio import join_room, emit, disconnect

import Project
from Project.settings import csrf
from Project.database import db
from user_app.models import User, Score
from ..models import Test, Room, Quiz
from Project.render_page import render_page

users = {}
devices = {}
user_devices = {}


def get_sid(username):
    for sid, name in users.items():
        if name == username:
            return sid
    
    return None

def room_get_result(room, author_name):
    room_get_result_data = {}
    
    ROOM = Room.query.filter_by(test_code= room).first()
    QUIZ_LIST = Quiz.query.filter_by(test_id= ROOM.test_id).all()
     
    users_list = ROOM.all_members.strip('|').split('||')
    USER_LIST = []
    UNREG_USER_LIST = []
    SCORE_LIST = []

    for user in users_list:
        if user:
            USER= User.query.filter_by(username=user).first()
            if USER:
                if USER.username != author_name:
                    USER_LIST.append(USER)
            else:
                if user != author_name:
                    UNREG_USER_LIST.append(user)

    SCORE_LIST= Score.query.filter_by(test_code=room).all()

    if USER_LIST:
        for user in USER_LIST:
            answers_list = []
            answers_str = ""
            correct_answers_list = []
            timers_list = []
            token_list = []

            for score in SCORE_LIST:
                if score.user_id == user.id:
                    timers_list = score.user_timers.split("|")
                    token_list = score.user_tokens.split("|")
                    answers_str = score.user_answer
            
            if answers_str:
                answers_list = answers_str.strip('|').split('||')

                for index, quiz in enumerate(QUIZ_LIST):
                    if answers_list[index] == "not_answer":
                        correct_answers_list.append(2)
                        continue
                    
                    if quiz.question_type  == "multiple_choice":
                        multi_choice_correct = quiz.correct_answer.split("%$№")
                        multi_choice_answer = answers_list[index].split("$$$")
                        sorted_multi_choice_correct = sorted(multi_choice_correct)
                        sorted_multi_choice_answer = sorted(multi_choice_answer)
                        if sorted_multi_choice_correct == sorted_multi_choice_answer:
                            correct_answers_list.append(1)
                        else:
                            correct_answers_list.append(0)
                    else:
                        if quiz.correct_answer == answers_list[index]:
                            correct_answers_list.append(1)
                        else:
                            correct_answers_list.append(0)

                room_get_result_data[user.username] = {
                    "correct_answers_list": correct_answers_list,
                    "timers_list": timers_list,
                    "token_list": token_list       
                }

    if(UNREG_USER_LIST):
        for user in UNREG_USER_LIST:
            answers_list = []
            answers_str = ""
            correct_answers_list = []
            timers_list = []
            token_list = []

            for score in SCORE_LIST:        
                if score.user_name == user:
                    timers_list = score.user_timers.split("|")
                    token_list = score.user_tokens.split("|")
                    answers_str = score.user_answer
            
            if answers_str:
                answers_list = answers_str.strip('|').split('||')

                for index, quiz in enumerate(QUIZ_LIST):
                    if answers_list[index] == "not_answer":
                        correct_answers_list.append(2)
                        continue

                    if quiz.question_type  == "multiple_choice":
                        multi_choice_correct = quiz.correct_answer.split("%$№")
                        multi_choice_answer = answers_list[index].split("$$$")
                        sorted_multi_choice_correct = sorted(multi_choice_correct)
                        sorted_multi_choice_answer = sorted(multi_choice_answer)
                        if sorted_multi_choice_correct == sorted_multi_choice_answer:
                            correct_answers_list.append(1)
                        else:
                            correct_answers_list.append(0)
                    else:
                        if quiz.correct_answer == answers_list[index]:
                            correct_answers_list.append(1)
                        else:
                            correct_answers_list.append(0)

                room_get_result_data[user] = {
                    "correct_answers_list": correct_answers_list,
                    "timers_list": timers_list,
                    "token_list": token_list         
                }

    BEST_SCORE = None
    best_accuracy = -1
    averega_accuracy = 0
    averega_score = 0

    for score in SCORE_LIST:
        averega_accuracy += score.accuracy
        if score.accuracy > best_accuracy:
            best_accuracy = score.accuracy
            BEST_SCORE = score

    WORST_SCORE = None
    worst_accuracy = 101

    for score in SCORE_LIST:
        if score.accuracy < worst_accuracy:
            worst_accuracy = score.accuracy
            WORST_SCORE = score

    if BEST_SCORE:
        averega_score = averega_accuracy // len(SCORE_LIST) 
    else:
        averega_score = 0

    best_score_data = {
        "user_name": BEST_SCORE.user_name if BEST_SCORE else "None",
        "accuracy": BEST_SCORE.accuracy if BEST_SCORE else 0,
    }

    worst_score_data = {
        "user_name": WORST_SCORE.user_name if WORST_SCORE else "None",
        "accuracy": WORST_SCORE.accuracy if WORST_SCORE else 0,
    }

    question_correct_count = [0] * len(QUIZ_LIST)

    for user_data in room_get_result_data.values():
        for index, answer in enumerate(user_data["correct_answers_list"]):
            if answer == 1:
                question_correct_count[index] += 1

    total_time_for_hardest_question = 0
    min_correct = min(question_correct_count)
    hardest_question_index = question_correct_count.index(min_correct)
    hardest_question = QUIZ_LIST[hardest_question_index]

    for user_data in room_get_result_data.values():
        timers = user_data["timers_list"]

        if timers and len(timers) > hardest_question_index and timers[hardest_question_index] and timers[hardest_question_index].strip():
            total_time_for_hardest_question += int(timers[hardest_question_index])

    # # Список пользователей, у которых есть результаты
    # user_results = []

    # for username, user_data in room_get_result_data.items():
    #     correct_count = sum(1 for a in user_data['correct_answers_list'] if a == 1)
    #     total_questions = len(user_data['correct_answers_list'])
    #     accuracy = (correct_count / total_questions) * 100 if total_questions > 0 else 0
    #     user_results.append({
    #         "user_name": username,
    #         "accuracy": accuracy
    #     })

    # # Лучший, худший и средний результат
    # if user_results:
    #     BEST_SCORE = max(user_results, key=lambda x: x['accuracy'])
    #     WORST_SCORE = min(user_results, key=lambda x: x['accuracy'])
    #     averega_score = sum(u['accuracy'] for u in user_results) / len(user_results)
    # else:
    #     BEST_SCORE = {"user_name": "None", "accuracy": 0}
    #     WORST_SCORE = {"user_name": "None", "accuracy": 0}
    #     averega_score = 0

    # best_score_data = {
    #     "user_name": BEST_SCORE["user_name"],
    #     "accuracy": BEST_SCORE["accuracy"]
    # }

    # worst_score_data = {
    #     "user_name": WORST_SCORE["user_name"],
    #     "accuracy": WORST_SCORE["accuracy"]
    # }

    # # Подсчет количества правильных ответов по каждому вопросу
    # question_correct_count = [0] * len(QUIZ_LIST)

    # for user_data in room_get_result_data.values():
    #     for index, answer in enumerate(user_data["correct_answers_list"]):
    #         if answer == 1:
    #             question_correct_count[index] += 1

    # # Самый сложный вопрос
    # total_time_for_hardest_question = 0
    # min_correct = min(question_correct_count)
    # hardest_question_index = question_correct_count.index(min_correct)
    # hardest_question = QUIZ_LIST[hardest_question_index]

    # # Суммарное время, потраченное на сложный вопрос
    # for user_data in room_get_result_data.values():
    #     timers = user_data.get("timers_list", [])
    #     if timers and len(timers) > hardest_question_index and timers[hardest_question_index].strip():
    #         total_time_for_hardest_question += int(timers[hardest_question_index])

    hardest_question_data = {
        "question_text": hardest_question.question_text,
        "correct_answers": min_correct,
        "total_time": total_time_for_hardest_question
    }

    return room_get_result_data, best_score_data, worst_score_data, hardest_question_data, averega_score


@Project.settings.socketio.on('join')
def handle_join(data):
    room = data['room']
    username = data['username']
    device_id = data["device_id"]
    user_sid = flask.request.sid

    if device_id in devices:
        emit("kicked", {"succses": "succses"}, to=user_sid)
        disconnect(sid=user_sid)

    users[user_sid] = username
    devices[device_id] = user_sid
    user_devices[username] = device_id
    join_room(room)

    test = Test.query.filter_by(test_code=room).first()
    ROOM = Room.query.filter_by(test_code=room).first()
    
    if not ROOM:
        NEW_ROOM = Room(
            test_id=test.id,
            test_code= room,
            user_list=f'|{username}|',
            author_name=username,
            active_test=False,
            all_members=""
        )
        db.session.add(NEW_ROOM)

    else:
        new_user = f'|{username}|'
        if new_user not in ROOM.user_list:
            ROOM.user_list += new_user
            if test and ROOM and username != test.author_name and username not in ROOM.all_members:
                ROOM.all_members += new_user

    db.session.commit()

@Project.settings.socketio.on('disconnect')
def handle_disconnect():
    username = users.pop(flask.request.sid, None)
    users.pop(flask.request.sid, None)

    if username:
        device_id = user_devices.pop(username, None)
        if device_id:
            devices.pop(device_id, None)
            ROOM = Room.query.filter(Room.user_list.like(f"%|{username}|%")).first()

            if ROOM and ROOM.user_list:
                ROOM.user_list = ROOM.user_list.replace(f"|{username}|", "")
                db.session.commit()

                emit('user_disconnected', {
                        'msg': f'{username} відключився',
                        "username": f"{username}"
                        }, 
                    to=ROOM.test_code)


@Project.settings.socketio.on('reconnect_user')
def handle_reconnect_user(data):
    author = data['author_name']
    author_sid = get_sid(author)

    emit("reconnect_ping", {"room": data["room"], "username": data["username"]}, room=author_sid)


@Project.settings.socketio.on('new_state')
def handle_new_state(data):
    new_state = data["new_state"]
    username = data["username"]
    user_sid = get_sid(username)
    print(new_state)
    emit("new_state", {"room": data["room"], "username": username, "new_state": new_state, "new_time": data["new_time"]}, room=user_sid)


@Project.settings.socketio.on('test_end')
def handle_clear_test_code(data):
    room = data['room']
    ROOM = Room.query.filter_by(test_code= room).first()

    if ROOM:
        db.session.delete(ROOM)

    TEST = Test.query.filter_by(test_code= room).first()
    if TEST:
        TEST.test_code = 0  
        
    db.session.commit()
    emit("test_end", room=room)

@Project.settings.socketio.on('test_reset')
def handle_clear_test_code(data):
    room = data['room']
    ROOM = Room.query.filter_by(test_code= room).first()

    if ROOM:
        db.session.delete(ROOM)

    TEST = Test.query.filter_by(test_code= room).first()
    if TEST:
        TEST.test_code = 0  
        
    db.session.commit()
    
@Project.settings.socketio.on('kick_user')
def handle_kick_user(data):
    username = data['user']
    room = data['room']
    
    # kick_sid = get_sid(username)
    device_id = user_devices.get(username)
    kick_sid = devices.get(device_id)

    if kick_sid:
        emit('kicked', room=kick_sid)
        disconnect(sid=kick_sid)

        users.pop(kick_sid, None)
        devices.pop(device_id, None)
        user_devices.pop(username, None)

        ROOM = Room.query.filter_by(test_code=room).first()
        ROOM.user_list = ROOM.user_list.replace(f"|{username}|", "")
        db.session.commit()
    else:
        print(f"Користувача {username} не знайдено серед підключених.")


@Project.settings.socketio.on('get_usernames')
def handle_send_usernames(data):
    room = data['room']
    author = data['author_name']
    author_sid = get_sid(author)

    ROOM = Room.query.filter_by(test_code=room).first()

    users_in_room = ROOM.user_list.split('|')
    clean_users_in_room = []
    for user in users_in_room:
        if user != author and user:
            clean_users_in_room.append(user)

    emit("get_usernames", clean_users_in_room, room=author_sid)


@Project.settings.socketio.on('user_answers')
def handle_user_answers(data):
    room = data["room"]
    user_name = data["username"]
    user_tokens = data["user_tokens"]
    tokens = 0
    new_token_list = None
    ROOM = Room.query.filter_by(test_code=room).first()
    TEST = Test.query.filter_by(id=ROOM.test_id).first()
    QUIZ_LIST = Quiz.query.filter_by(test_id=ROOM.test_id).all()

    
    user_answers = data["user_answers"].split("|")
    user_answers_list = []
    for answer in user_answers:
        if answer != "":
            user_answers_list.append(answer)

    for token in user_tokens.split("|"):
        if token:
            tokens += int(token)
    
    number_of_correct_answers = 0
    user_tokens = user_tokens.split("|")

    for index in range(len(QUIZ_LIST)):
        user = user_answers_list[index].split("$$$")
        quiz = QUIZ_LIST[index].correct_answer.split("%$№")
        if sorted(user) == sorted(quiz):
            number_of_correct_answers += 1
        else:
            user_tokens[index] = 0

    for token in user_tokens:
        if new_token_list == None:
            new_token_list = f"{token}"
        else:
            new_token_list += f"|{token}"

    accuracy = number_of_correct_answers / len(QUIZ_LIST) * 100
    accuracy = int(accuracy)
    USER = User.query.filter_by(username= user_name).first()

    SCORE = Score(
        user_answer=data["user_answers"],
        user_timers=data["user_timers"],
        user_tokens=new_token_list,
        accuracy=accuracy,
        test_id=TEST.id,
        date_complete=datetime.date.today(),
        time_complete=datetime.datetime.now().strftime("%H:%M:%S"),
        user_id=USER.id if USER else None,
        user_name=user_name,
        test_code=room
    )

    if USER:
        if USER.tokens:
            USER.tokens = int(USER.tokens) + tokens
        else:
            USER.tokens = tokens

    db.session.add(SCORE)
    db.session.commit() 

    author_sid = get_sid(TEST.author_name)
    if author_sid:
        room_get_result_data, best_score_data, worst_score_data, hardest_question_data, averega_score = room_get_result(room, TEST.author_name)

        user_result = room_get_result_data.get(user_name)
        if user_result:      
            emit("author_receive_new_result", {"username": user_name, "user_result": user_result}, room=author_sid)


@Project.settings.socketio.on('user_answer')
def handle_user_answer(data):
    author_name = data['author_name']
    username = data['username']
    answer = data['answer']
    
    author_sid = get_sid(author_name)

    emit("author_receive_answer", {"username": username, "answer": answer}, room=author_sid)


@Project.settings.socketio.on('get_room_size')
def handle_get_room_size(data):
    room = data["room"]
    author_name = data["author_name"]

    author_sid = get_sid(author_name)
    
    count_users = 0
    ROOM = Room.query.filter_by(test_code=room).first()
    user_str = ROOM.user_list
    
    for user in user_str.split('|'):
        if user != "":
            count_users = count_users + 1

    count_users = count_users - 1
    
    emit("recieve_count_users", {
        "room": room,
        "countUsers": count_users
        }
    , to= author_sid)


@Project.settings.socketio.on('author_start_test')
def handle_start_test(data):
    room = data['room']
    test = Test.query.filter_by(test_code=room).first()
    
    ROOM = Room.query.filter_by(test_code=room).first()
    user_list = ROOM.user_list.replace(f"|{ROOM.author_name}|", "")

    if user_list:
        ROOM.active_test = True
        db.session.commit()
        
        emit("start_test", {
            "room": room,
            "test_id": test.id,
            "author_name": test.author_name
            }
        , to=room)


@Project.settings.socketio.on('message_to_chat')
def handle_message_to_chat(data):
    emit("listening_to_messages", f"{data['username']}: {data['message']}", include_self=False, to=data['room'])


@Project.settings.socketio.on('new_user')
def handle_new_user(data):
    room = data['room']
    username = data['username']

    ROOM = Room.query.filter_by(test_code=room).first()
    users_list = ROOM.user_list

    emit("create_user_block", {"username": username, "user_ip": data["user_ip"], "users_list": users_list}, to=room)


@Project.settings.socketio.on('new_user_admin')
def handle_new_user_admin(data):
    author_name = data["author_name"]
    compound = data.get("compound", 0)

    author_sid = get_sid(author_name)
    emit("new_user_admin", {"username": data["username"], "ip": data["ip"], "compound": compound}, to=author_sid)


@Project.settings.socketio.on('next_question')
def handle_next_question(data):
    emit("next_question", f"Next question in {data['room']} author {data['author_name']}", include_self=False, to=data['room'])


@Project.settings.socketio.on('stop_test')
def handle_stop_test(data):
    emit("result_test", f"Stop test {data['room']} result_test page author {data['author_name']}", include_self=False, to=data['room'])


@Project.settings.socketio.on('end_test')
def handle_end_test(data):
    room = data["room"]
    emit("kicked", room, include_self=False, to=room)


@Project.settings.socketio.on('room_get_result')
def handle_room_get_result(data):
    user_sid = get_sid(data["username"])
    room_get_result_data, best_score_data, worst_score_data, hardest_question_data, averega_score = room_get_result(data["room"], data["author_name"])
   
    emit("room_get_result_data", {
        "room_get_result_data": room_get_result_data,
        "best_score_data": best_score_data,
        "worst_score_data": worst_score_data,
        "hardest_question_data": hardest_question_data,
        "averega_score": averega_score
    }, to= user_sid)


@Project.settings.socketio.on('plus_time')
def handle_plus_time(data):
    emit("plus_time", to=data['room'])


@Project.settings.socketio.on('change_time')
def handle_change_time(data):
    emit("change_time", to=data['room'])

@Project.settings.socketio.on('user_leave')
def handle_user_leave(data):
    emit("user_leave", {"leave_user": data["leave_user"]}, to=data['room'])

def excel_table(username, author_name, result_data, best_score_data, worst_score_data, hardest_question_data, test_code):
    """
    result_data: dict
    best_score_data: dict {'user_name': str, 'accuracy': int}
    worst_score_data: dict {'user_name': str, 'accuracy': int}
    hardest_question_data: dict {'question_text': str, 'correct_answers': int, "total_time": int}
    average_score: int | float
    """

    if not result_data:
        return 

    first_user = next(iter(result_data))
    total_questions = len(result_data[first_user]["correct_answers_list"])

    # Заголовок
    header_row = ["Учень"]
    for i in range(1, total_questions + 1):
        header_row.append(f"Q{i}")
    header_row.append("Точність")

    table = [header_row]

    total_accuracy = 0
    users_count = 0

    for user, data in result_data.items():
        answers = data["correct_answers_list"]

        correct = sum(1 for a in answers if a == 1)
        accuracy_number = correct / len(answers)
        accuracy = f"{accuracy_number * 100:.1f}%"

        total_accuracy += accuracy_number
        users_count += 1

        row = [user]

        for a in answers:
            if a == 1:
                row.append("✅")
            elif a == 0:
                row.append("❌")
            else:
                row.append("-")

        row.append(accuracy)
        table.append(row)

    average_accuracy = f"{(total_accuracy / users_count) * 100:.1f}%"

    # Пустая строка
    table.append([])

    # Итоги
    table.append(["Найкращий результат", best_score_data["user_name"], f'{best_score_data["accuracy"]}%'])
    table.append(["Гірший результат", worst_score_data["user_name"], f'{worst_score_data["accuracy"]}%'])
    table.append(["Середній результат", average_accuracy])
    table.append(["Найскладніше питання", hardest_question_data["question_text"]])
    table.append(["Кількість правильних відповідей", hardest_question_data["correct_answers"]])
    table.append(["Загальний час", hardest_question_data["total_time"]])



    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    for row in table:
        ws.append(row)
    #сохранять с f строкой 


    # ---------------- ЛИСТ С ГРАФИКОМ ----------------

    accuracy_ws = wb.create_sheet("Accuracy_Line")

    labels = []
    accuracy_numbers = []

    first_user = next(iter(result_data))
    total_questions = len(result_data[first_user]["correct_answers_list"])

    # если нужен конкретный user (аналог userName)
    selected_user = None  # сюда можешь передать username если нужно

    if selected_user and selected_user in result_data:
        answers = result_data[selected_user]["correct_answers_list"]
        accuracy_numbers = [100 if a == 1 else 0 for a in answers]
        labels = [f"Q{i+1}" for i in range(total_questions)]
    else:
        labels = [f"Q{i+1}" for i in range(total_questions)]
        for i in range(total_questions):
            correct = 0
            for user in result_data:
                if result_data[user]["correct_answers_list"][i] == 1:
                    correct += 1
            percent = (correct / len(result_data)) * 100
            accuracy_numbers.append(percent)

    accuracy_ws.append(["Питання", "Точність (%)"])

    for i in range(total_questions):
        accuracy_ws.append([labels[i], accuracy_numbers[i]])

    line_chart = LineChart()
    line_chart.title = "Точність відповідей (%)"
    line_chart.y_axis.title = "Точність (%)"
    line_chart.x_axis.title = "Номер питання"

    data = Reference(accuracy_ws, min_col=2, min_row=1,
                    max_row=total_questions + 1)
    cats = Reference(accuracy_ws, min_col=1, min_row=2,
                    max_row=total_questions + 1)

    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(cats)

    accuracy_ws.add_chart(line_chart, "E2")

    # ---------------- CORRECT / WRONG BAR CHART ----------------

    bar_ws = wb.create_sheet("Correct_Wrong_Bar")

    correct_counts = [0] * total_questions
    wrong_counts = [0] * total_questions

    for user in result_data:
        answers = result_data[user]["correct_answers_list"]
        for i in range(total_questions):
            if answers[i] == 1:
                correct_counts[i] += 1
            else:
                wrong_counts[i] += 1

    wrong_counts = [-x for x in wrong_counts]

    bar_ws.append(["Питання", "Правильно", "Неправильно"])

    for i in range(total_questions):
        bar_ws.append([f"Q{i+1}", correct_counts[i], wrong_counts[i]])

    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Правильно / Неправильно"
    bar_chart.y_axis.title = "Кількість користувачів"
    bar_chart.x_axis.title = "Номер питання"

    data = Reference(bar_ws, min_col=2, max_col=3,
                    min_row=1, max_row=total_questions + 1)
    cats = Reference(bar_ws, min_col=1,
                    min_row=2, max_row=total_questions + 1)

    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)

    # динамический масштаб как в JS
    max_correct = max(correct_counts)
    max_wrong = max([-x for x in wrong_counts])

    bar_chart.y_axis.scaling.min = -math.ceil(max_wrong * 1.2)
    bar_chart.y_axis.scaling.max = math.ceil(max_correct * 1.2)

    bar_ws.add_chart(bar_chart, "E2")


    # ---------------- PIE CHART ----------------

    pie_ws = wb.create_sheet("User_Pie")

    users = list(result_data.keys())
    total_slices = len(users) * total_questions

    correct_total = 0
    values = []

    for user in users:
        correct_count = sum(1 for a in result_data[user]["correct_answers_list"] if a == 1)
        correct_total += correct_count
        values.append(correct_count)

    remaining = total_slices - correct_total

    labels = users.copy()

    if remaining > 0:
        values.append(remaining)
        labels.append("Неправильні / пропущені")

    pie_ws.append(["Користувач", "Кількість"])

    for i in range(len(values)):
        pie_ws.append([labels[i], values[i]])

    pie_chart = PieChart()
    pie_chart.title = "Розподіл правильних відповідей"

    data = Reference(pie_ws, min_col=2, min_row=1,
                    max_row=len(values) + 1)
    cats = Reference(pie_ws, min_col=1, min_row=2,
                    max_row=len(values) + 1)

    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(cats)

    pie_ws.add_chart(pie_chart, "E2")


    # ---------------- TIME / TOKEN LINE CHART ----------------

    time_ws = wb.create_sheet("Question_Time")

    total_timers = [0] * total_questions
    type_value = "time"  # или "token"

    for user in result_data:
        if type_value == "token":
            type_list = result_data[user]["token_list"]
        else:
            type_list = result_data[user]["timers_list"]

        for i in range(total_questions):
            if type_list[i]:
                total_timers[i] += int(type_list[i])

    time_ws.append(["Питання", "Суммарне значення"])

    for i in range(total_questions):
        time_ws.append([f"Q{i+1}", total_timers[i]])

    time_chart = LineChart()
    time_chart.title = "Суммарне значення по питанням"
    time_chart.y_axis.title = "Значення"
    time_chart.x_axis.title = "Номер питання"

    data = Reference(time_ws, min_col=2, min_row=1,
                    max_row=total_questions + 1)
    cats = Reference(time_ws, min_col=1, min_row=2,
                    max_row=total_questions + 1)

    time_chart.add_data(data, titles_from_data=True)
    time_chart.set_categories(cats)

    time_ws.add_chart(time_chart, "E2")


    # ---------------- DOUGHNUT CHART ----------------

    doughnut_ws = wb.create_sheet("Overall_Doughnut")

    total_answers = len(result_data) * total_questions
    correct_count = sum(
        1 for user in result_data
        for a in result_data[user]["correct_answers_list"]
        if a == 1
    )

    correct_percent = (correct_count / total_answers) * 100
    incorrect_percent = 100 - correct_percent

    doughnut_ws.append(["Тип", "Відсоток"])
    doughnut_ws.append(["Правильні (%)", correct_percent])
    doughnut_ws.append(["Неправильні (%)", incorrect_percent])

    doughnut_chart = DoughnutChart()
    doughnut_chart.title = "Загальна точність (%)"

    data = Reference(doughnut_ws, min_col=2, min_row=1, max_row=3)
    cats = Reference(doughnut_ws, min_col=1, min_row=2, max_row=3)

    doughnut_chart.add_data(data, titles_from_data=True)
    doughnut_chart.set_categories(cats)

    doughnut_ws.add_chart(doughnut_chart, "E2")


    # ---------------- СОХРАНЕНИЕ ----------------
    wb.save(f"results{test_code}.xlsx")

@csrf.exempt
@render_page(template_name = 'room.html')
def render_room(test_code):
    list_answers = []
    list_quiz = []

    test = Test.query.filter_by(test_code=test_code).first()

    if not test:
        return flask.redirect("/")

    quizzes_list = Quiz.query.filter_by(test_id= test.id).all()

    for quiz in quizzes_list:
        if quiz.question_type == "input":
            list_answers.append(quiz.correct_answer)
            list_quiz.append(quiz.dict())
        else :
            list_answers.append(quiz.answer_options.split("%$№"))
            list_quiz.append(quiz.dict()) 
    
    if flask.request.method == "POST":
        room_get_result_data, best_score_data, worst_score_data, hardest_question_data, averega_score = room_get_result(test_code, test.author_name)
        base_dir = os.path.dirname(
            os.path.dirname(
                os.path.dirname(os.path.abspath(__file__))
            )
        )
        file_path = os.path.join(base_dir, f"results{test_code}.xlsx")
        excel_table(
            username="teacher",
            author_name="admin",
            result_data=room_get_result_data,
            best_score_data=best_score_data,
            worst_score_data=worst_score_data,
            hardest_question_data=hardest_question_data,
            test_code=test_code
        )
        
        return_data = io.BytesIO()
        with open(file_path, 'rb') as f:
            return_data.write(f.read())
        return_data.seek(0)

        os.remove(file_path)

        return flask.send_file(
            return_data,
            as_attachment=True,
            download_name="results.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    return {
        "test": test,
        "list_quiz": list_quiz,
        "list_answers": list_answers,
    }


def delete_code(test_id):
    test = Test.query.filter_by(id= test_id).first()
    test.test_code = 0
    Project.database.db.session.commit()

    return flask.redirect("/quizzes/")