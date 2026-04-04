import math

from openpyxl import Workbook
from openpyxl.chart import LineChart, PieChart, DoughnutChart, Reference, BarChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import PatternFillProperties
from openpyxl.drawing.colors import ColorChoice


from user_app.models import User, Score
from ..models import Room, Quiz

def clean_data(data):
    cleaned = {}
    for user, values in data.items():
        if not values:
            continue

        cleaned[user] = {
            "correct_answers_list": values.get("correct_answers_list") or [],
            "timers_list": values.get("timers_list") or [],
            "token_list": values.get("token_list") or []
        }
    return cleaned

def room_get_result(room, author_name):
    room_get_result_data = {}

    print(room, author_name)
    
    ROOM = Room.query.filter_by(test_code= room).first()
    print("ROOM найден:", ROOM)

    if ROOM:
        QUIZ_LIST = Quiz.query.filter_by(test_id=ROOM.test_id).all()
        print("Количество вопросов в QUIZ_LIST:", len(QUIZ_LIST))
    else:
        print("Room с таким кодом теста не найден")

    if not ROOM: 
        return {}, {}, {}, {}, 0
     
    QUIZ_LIST = Quiz.query.filter_by(test_id= ROOM.test_id).all()

    if not QUIZ_LIST:
        return {}, {}, {}, {}, 0
     
    users_list = ROOM.all_members.strip('|').split('||')
    USER_LIST = []
    UNREG_USER_LIST = []
    SCORE_LIST = []

    for user in users_list:
        if user:
            USER= User.query.filter_by(username=user).first()
            if USER:
                if USER.username != author_name:
                    USER_LIST.append(USER)
            else:
                if user != author_name:
                    UNREG_USER_LIST.append(user)

    SCORE_LIST= Score.query.filter_by(test_code=room).all()

    if USER_LIST:
        for user in USER_LIST:
            answers_list = []
            answers_str = ""
            correct_answers_list = []
            timers_list = []
            token_list = []

            for score in SCORE_LIST:
                if score.user_id == user.id:
                    timers_list = score.user_timers.split("|") if score.user_timers else []
                    token_list = score.user_tokens.split("|") if score.user_tokens else []
                    answers_str = score.user_answer
            
            if answers_str:
                # print(answers_str)
                answers_list = answers_str.strip('|').split('||')
                # print(answers_str.strip('|'))

                # print(QUIZ_LIST)
                # print(answers_list)
                for index, quiz in enumerate(QUIZ_LIST):
                    ans = answers_list[index] if index < len(answers_list) else "not_answer"

                    if ans == "not_answer":
                        correct_answers_list.append(2)
                        continue
                    
                    if quiz.question_type  == "multiple_choice":
                        multi_choice_correct = quiz.correct_answer.split("%$№")
                        multi_choice_answer = ans.split("$$$")
                        sorted_multi_choice_correct = sorted(multi_choice_correct)
                        sorted_multi_choice_answer = sorted(multi_choice_answer)

                        if sorted_multi_choice_correct == sorted_multi_choice_answer:
                            correct_answers_list.append(1)
                        else:
                            correct_answers_list.append(0)
                    else:
                        if quiz.correct_answer == ans:
                            correct_answers_list.append(1)
                        else:
                            correct_answers_list.append(0)

                room_get_result_data[user.username] = {
                    "correct_answers_list": correct_answers_list,
                    "timers_list": timers_list,
                    "token_list": token_list       
                }

    if(UNREG_USER_LIST):
        for user in UNREG_USER_LIST:
            answers_list = []
            answers_str = ""
            correct_answers_list = []
            timers_list = []
            token_list = []

            for score in SCORE_LIST:        
                if score.user_name == user:
                    timers_list = score.user_timers.split("|") if score.user_timers else []
                    token_list = score.user_tokens.split("|") if score.user_tokens else []
                    answers_str = score.user_answer
            
            if answers_str:
                answers_list = answers_str.strip('|').split('||')

                for index, quiz in enumerate(QUIZ_LIST):
                    ans = answers_list[index] if index < len(answers_list) else "not_answer"

                    if ans == "not_answer":
                        correct_answers_list.append(2)
                        continue

                    if quiz.question_type  == "multiple_choice":
                        multi_choice_correct = quiz.correct_answer.split("%$№")
                        multi_choice_answer = ans.split("$$$")
                        sorted_multi_choice_correct = sorted(multi_choice_correct)
                        sorted_multi_choice_answer = sorted(multi_choice_answer)
                        if sorted_multi_choice_correct == sorted_multi_choice_answer:
                            correct_answers_list.append(1)
                        else:
                            correct_answers_list.append(0)
                    else:
                        if quiz.correct_answer == ans:
                            correct_answers_list.append(1)
                        else:
                            correct_answers_list.append(0)

                room_get_result_data[user] = {
                    "correct_answers_list": correct_answers_list,
                    "timers_list": timers_list,
                    "token_list": token_list         
                }

    BEST_SCORE = None
    best_accuracy = -1
    averega_accuracy = 0
    averega_score = 0

    for score in SCORE_LIST:
        averega_accuracy += score.accuracy
        if score.accuracy > best_accuracy:
            best_accuracy = score.accuracy
            BEST_SCORE = score

    WORST_SCORE = None
    worst_accuracy = 101

    for score in SCORE_LIST:
        if score.accuracy < worst_accuracy:
            worst_accuracy = score.accuracy
            WORST_SCORE = score

    if BEST_SCORE:
        averega_score = averega_accuracy // len(SCORE_LIST) 
    else:
        averega_score = 0

    best_score_data = {
        "user_name": BEST_SCORE.user_name if BEST_SCORE else "None",
        "accuracy": BEST_SCORE.accuracy if BEST_SCORE else 0,
    }

    worst_score_data = {
        "user_name": WORST_SCORE.user_name if WORST_SCORE else "None",
        "accuracy": WORST_SCORE.accuracy if WORST_SCORE else 0,
    }

    question_correct_count = [0] * len(QUIZ_LIST)

    for user_data in room_get_result_data.values():
        for index, answer in enumerate(user_data["correct_answers_list"]):
            if answer == 1:
                question_correct_count[index] += 1

    total_time_for_hardest_question = 0
    min_correct = min(question_correct_count)
    hardest_question_index = question_correct_count.index(min_correct)
    hardest_question = QUIZ_LIST[hardest_question_index]

    for user_data in room_get_result_data.values():
        timers = user_data["timers_list"]

        if timers and len(timers) > hardest_question_index and timers[hardest_question_index] and timers[hardest_question_index].strip():
            total_time_for_hardest_question += int(timers[hardest_question_index])

    hardest_question_data = {
        "question_text": hardest_question.question_text,
        "correct_answers": min_correct,
        "total_time": int(total_time_for_hardest_question),
        "hardest_question_index": hardest_question_index
    }

    print("USER_LIST (зарегистрированные пользователи):", [u.username for u in USER_LIST])
    print("UNREG_USER_LIST (незарегистрированные пользователи):", UNREG_USER_LIST)
    print("SCORE_LIST:", SCORE_LIST)

    for username, data in room_get_result_data.items():
        print(f"Результаты пользователя {username}: {data}")

    if QUIZ_LIST:
        print("Самый сложный вопрос:", hardest_question.question_text, "индекс:", hardest_question_data.get("hardest_question_index"))
    else:
        print("Нет вопросов для анализа hardest_question")

    return room_get_result_data, best_score_data, worst_score_data, hardest_question_data, averega_score


def excel_table(username, author_name, result_data, best_score_data, worst_score_data, hardest_question_data, test_code):
    """
    result_data: dict
    best_score_data: dict {'user_name': str, 'accuracy': int}
    worst_score_data: dict {'user_name': str, 'accuracy': int}
    hardest_question_data: dict {'question_text': str, 'correct_answers': int, "total_time": int}
    average_score: int | float
    """

    if not result_data:
        return 

    first_user = next(iter(result_data))
    total_questions = len(result_data[first_user]["correct_answers_list"])

    # Заголовок
    header_row = ["Учень"]
    for i in range(1, total_questions + 1):
        header_row.append(f"Q{i}")
    header_row.append("Точність")

    table = [header_row]

    total_accuracy = 0
    users_count = 0

    for user, data in result_data.items():
        answers = data["correct_answers_list"]

        correct = sum(1 for a in answers if a == 1)
        accuracy_number = correct / len(answers)
        accuracy = f"{accuracy_number * 100:.1f}%"

        total_accuracy += accuracy_number
        users_count += 1

        row = [user]

        for a in answers:
            if a == 1:
                row.append("✅")
            elif a == 0:
                row.append("❌")
            else:
                row.append("-")

        row.append(accuracy)
        table.append(row)

    average_accuracy = f"{(total_accuracy / users_count) * 100:.1f}%"

    # Пустая строка
    table.append([])

    # Итоги
    table.append(["Найкращий результат", best_score_data["user_name"], f'{best_score_data["accuracy"]}%'])
    table.append(["Гірший результат", worst_score_data["user_name"], f'{worst_score_data["accuracy"]}%'])
    table.append(["Середній результат", average_accuracy])
    table.append(["Найскладніше питання", hardest_question_data["question_text"]])
    table.append(["Кількість правильних відповідей", hardest_question_data["correct_answers"]])
    table.append(["Загальний час", hardest_question_data["total_time"]])



    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    for row in table:
        ws.append(row)
    #сохранять с f строкой 


    # ---------------- ЛИСТ С ГРАФИКОМ ----------------

    accuracy_ws = wb.create_sheet("Accuracy_Line")

    labels = []
    accuracy_numbers = []

    first_user = next(iter(result_data))
    total_questions = len(result_data[first_user]["correct_answers_list"])

    # если нужен конкретный user (аналог userName)
    selected_user = None  # сюда можешь передать username если нужно

    if selected_user and selected_user in result_data:
        answers = result_data[selected_user]["correct_answers_list"]
        accuracy_numbers = [100 if a == 1 else 0 for a in answers]
        labels = [f"Q{i+1}" for i in range(total_questions)]
    else:
        labels = [f"Q{i+1}" for i in range(total_questions)]
        for i in range(total_questions):
            correct = 0
            for user in result_data:
                if result_data[user]["correct_answers_list"][i] == 1:
                    correct += 1
            percent = (correct / len(result_data)) * 100
            accuracy_numbers.append(percent)

    accuracy_ws.append(["Питання", "Точність (%)"])

    for i in range(total_questions):
        accuracy_ws.append([labels[i], accuracy_numbers[i]])

    line_chart = LineChart()
    line_chart.title = "Точність відповідей (%)"
    line_chart.y_axis.title = "Точність (%)"
    line_chart.x_axis.title = "Номер питання"

    data = Reference(accuracy_ws, min_col=2, min_row=1,
                    max_row=total_questions + 1)
    cats = Reference(accuracy_ws, min_col=1, min_row=2,
                    max_row=total_questions + 1)

    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(cats)

    accuracy_ws.add_chart(line_chart, "C1")

    # ---------------- CORRECT / WRONG BAR CHART ----------------

    bar_ws = wb.create_sheet("Correct_Wrong_Bar")

    correct_counts = [0] * total_questions
    wrong_counts = [0] * total_questions

    for user in result_data:
        answers = result_data[user]["correct_answers_list"]
        for i in range(total_questions):
            if answers[i] == 1:
                correct_counts[i] += 1
            else:
                wrong_counts[i] += 1

    wrong_counts = [-x for x in wrong_counts]

    bar_ws.append(["Питання", "Правильно", "Неправильно"])

    for i in range(total_questions):
        bar_ws.append([f"Q{i+1}", correct_counts[i], abs(wrong_counts[i])])

    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Правильно / Неправильно"
    bar_chart.y_axis.title = "Кількість користувачів"
    bar_chart.x_axis.title = "Номер питання"

    data = Reference(bar_ws, min_col=2, max_col=3,
                    min_row=1, max_row=total_questions + 1)
    cats = Reference(bar_ws, min_col=1,
                    min_row=2, max_row=total_questions + 1)

    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)

    # динамический масштаб как в JS
    max_correct = max(correct_counts)
    max_wrong = max([-x for x in wrong_counts])

    bar_chart.y_axis.scaling.min = -math.ceil(max_wrong * 1.2)
    bar_chart.y_axis.scaling.max = math.ceil(max_correct * 1.2)

    bar_ws.add_chart(bar_chart, "D1")


    # ---------------- PIE CHART ----------------

    pie_ws = wb.create_sheet("User_Pie")

    users = list(result_data.keys())
    total_slices = len(users) * total_questions

    correct_total = 0
    values = []

    for user in users:
        correct_count = sum(1 for a in result_data[user]["correct_answers_list"] if a == 1)
        correct_total += correct_count
        values.append(correct_count)

    remaining = total_slices - correct_total

    labels = users.copy()

    if remaining > 0:
        values.append(remaining)
        labels.append("Неправильні / пропущені")

    pie_ws.append(["Користувач", "Кількість"])

    for i in range(len(values)):
        pie_ws.append([labels[i], values[i]])

    pie_chart = PieChart()
    pie_chart.title = "Розподіл правильних відповідей"

    data = Reference(pie_ws, min_col=2, min_row=1,
                    max_row=len(values) + 1)
    cats = Reference(pie_ws, min_col=1, min_row=2,
                    max_row=len(values) + 1)

    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(cats)

    pie_ws.add_chart(pie_chart, "C1")


    # ---------------- TIME / TOKEN LINE CHART ----------------

    time_ws = wb.create_sheet("Question_Time")

    total_timers = [0] * total_questions
    type_value = "time"  # или "token"

    for user in result_data:
        if type_value == "token":
            type_list = result_data[user]["token_list"]
        else:
            type_list = result_data[user]["timers_list"]

        for i in range(total_questions):
            total_timers[i] += float(type_list[i])

    time_ws.append(["Питання", "Витрачено часу на запитання (сек)"])

    for i in range(total_questions):
        time_ws.append([f"Q{i+1}", total_timers[i]])

    time_chart = LineChart()
    time_chart.title = "Сумарний час на питання (сек)"
    time_chart.y_axis.title = "Значення"
    time_chart.x_axis.title = "Номер питання"

    data = Reference(time_ws, min_col=2, min_row=1,
                    max_row=total_questions + 1)
    cats = Reference(time_ws, min_col=1, min_row=2,
                    max_row=total_questions + 1)

    time_chart.add_data(data, titles_from_data=True)
    time_chart.set_categories(cats)

    time_ws.add_chart(time_chart, "C1")


    # ---------------- DOUGHNUT CHART ----------------

    doughnut_ws = wb.create_sheet("Overall_Doughnut")

    total_answers = len(result_data) * total_questions
    correct_count = sum(
        1 for user in result_data
        for a in result_data[user]["correct_answers_list"]
        if a == 1
    )

    correct_percent = (correct_count / total_answers) * 100
    incorrect_percent = 100 - correct_percent

    doughnut_ws.append(["Тип", "Відсоток"])
    doughnut_ws.append(["Правильні (%)", round(correct_percent, 1)])
    doughnut_ws.append(["Неправильні (%)", round(incorrect_percent, 1)])

    doughnut_chart = DoughnutChart()
    doughnut_chart.title = "Загальна точнисть правильних відповідей (%)"

    data = Reference(doughnut_ws, min_col=2, min_row=1, max_row=3)
    cats = Reference(doughnut_ws, min_col=1, min_row=2, max_row=3)

    doughnut_chart.add_data(data, titles_from_data=True)
    doughnut_chart.set_categories(cats)

    doughnut_ws.add_chart(doughnut_chart, "C1")


    # ---------------- СОХРАНЕНИЕ ----------------
    wb.save(f"results{test_code}.xlsx")